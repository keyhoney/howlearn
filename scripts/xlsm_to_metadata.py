"""
XLSM -> problem metadata JSON 변환기.

사용 예시:
  python scripts/xlsm_to_metadata.py --xlsm "C:\\path\\problem-metadata.xlsm"

기본 시트 이름:
  - problems
  - essay-problems

출력 경로:
  - src/content/problems/_metadata.json
  - src/content/essay-problems/_metadata.json
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROBLEMS_OUT = ROOT / "src" / "content" / "problems" / "_metadata.json"
ESSAY_OUT = ROOT / "src" / "content" / "essay-problems" / "_metadata.json"


@dataclass(frozen=True)
class SheetSpec:
    name: str
    required_columns: tuple[str, ...]
    output_path: Path


PROBLEMS_SPEC = SheetSpec(
    name="problems",
    required_columns=(
        "id",
        "source",
        "year",
        "month",
        "examType",
        "subject",
        "chapter",
        "subChapter",
        "concept",
        "difficulty",
        "answerType",
        "answer",
    ),
    output_path=PROBLEMS_OUT,
)

ESSAY_SPEC = SheetSpec(
    name="essay-problems",
    required_columns=(
        "id",
        "source",
        "year",
        "examType",
        "difficulty",
        "university",
    ),
    output_path=ESSAY_OUT,
)


def _to_int(value: Any, *, field: str, row_number: int, sheet_name: str) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"[{sheet_name}:{row_number}] {field} 값이 비어 있습니다.")
    try:
        return int(float(str(value).strip()))
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"[{sheet_name}:{row_number}] {field}는 정수여야 합니다: {value}") from exc


def _to_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _to_list(value: Any) -> list[str]:
    raw = _to_str(value)
    if not raw:
        return []
    return [item.strip() for item in raw.split(",") if item.strip()]


def _validate_headers(sheet, spec: SheetSpec) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError(f"[{spec.name}] 헤더 행이 없습니다.")

    headers_raw: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _to_str(cell)
        if key:
            headers_raw[key] = idx

    # 헤더는 대소문자를 구분하지 않고 처리한다.
    # 반환 맵에는 실제 키 + 소문자 키 + 자주 쓰는 camelCase 별칭을 함께 넣어
    # 호출부가 기존 키명을 그대로 써도 동작하도록 유지한다.
    headers: dict[str, int] = {}
    for key, idx in headers_raw.items():
        lower = key.lower()
        headers[key] = idx
        headers[lower] = idx
        if lower == "subchapter":
            headers["subChapter"] = idx
        elif lower == "examyear":
            headers["examYear"] = idx
        elif lower == "examtype":
            headers["examType"] = idx
        elif lower == "answertype":
            headers["answerType"] = idx
        elif lower == "relatedproblemids":
            headers["relatedProblemIds"] = idx

    missing = [col for col in spec.required_columns if col.lower() not in headers]
    if missing:
        raise ValueError(f"[{spec.name}] 필수 컬럼 누락: {', '.join(missing)}")
    return headers


def _build_problems_entry(row: tuple[Any, ...], headers: dict[str, int], row_number: int) -> tuple[str, dict[str, Any]] | None:
    problem_id = _to_str(row[headers["id"]]) if headers["id"] < len(row) else ""
    if not problem_id:
        return None

    item: dict[str, Any] = {
        "source": _to_str(row[headers["source"]]),
        "year": _to_int(row[headers["year"]], field="year", row_number=row_number, sheet_name="problems"),
        "month": _to_int(row[headers["month"]], field="month", row_number=row_number, sheet_name="problems"),
        "examType": _to_str(row[headers["examType"]]),
        "subject": _to_str(row[headers["subject"]]),
        "chapter": _to_str(row[headers["chapter"]]),
        "subChapter": _to_str(row[headers["subChapter"]]),
        "concept": _to_str(row[headers["concept"]]),
        "difficulty": _to_int(row[headers["difficulty"]], field="difficulty", row_number=row_number, sheet_name="problems"),
        "answerType": _to_str(row[headers["answerType"]]),
        "answer": _to_int(row[headers["answer"]], field="answer", row_number=row_number, sheet_name="problems"),
    }
    return problem_id, item


def _build_essay_entry(row: tuple[Any, ...], headers: dict[str, int], row_number: int) -> tuple[str, dict[str, Any]] | None:
    essay_id = _to_str(row[headers["id"]]) if headers["id"] < len(row) else ""
    if not essay_id:
        return None

    exam_type = _to_str(row[headers["examType"]]) or "논술"
    item: dict[str, Any] = {
        "source": _to_str(row[headers["source"]]),
        "year": _to_int(row[headers["year"]], field="year", row_number=row_number, sheet_name="essay-problems"),
        "examType": exam_type,
        "difficulty": _to_int(row[headers["difficulty"]], field="difficulty", row_number=row_number, sheet_name="essay-problems"),
        "university": _to_str(row[headers["university"]]),
    }

    if "examYear" in headers and headers["examYear"] < len(row):
        exam_year_raw = row[headers["examYear"]]
        if _to_str(exam_year_raw):
            item["examYear"] = _to_int(exam_year_raw, field="examYear", row_number=row_number, sheet_name="essay-problems")

    if "tags" in headers and headers["tags"] < len(row):
        item["tags"] = _to_list(row[headers["tags"]])
    else:
        item["tags"] = []

    if "relatedProblemIds" in headers and headers["relatedProblemIds"] < len(row):
        item["relatedProblemIds"] = _to_list(row[headers["relatedProblemIds"]])
    else:
        item["relatedProblemIds"] = []

    return essay_id, item


def _read_sheet(sheet, spec: SheetSpec) -> dict[str, dict[str, Any]]:
    headers = _validate_headers(sheet, spec)
    out: dict[str, dict[str, Any]] = {}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_to_str(v) for v in row):
            continue

        if spec.name == "problems":
            built = _build_problems_entry(row, headers, row_number)
        else:
            built = _build_essay_entry(row, headers, row_number)

        if built is None:
            continue
        key, value = built
        if key in out:
            raise ValueError(f"[{spec.name}:{row_number}] 중복 id: {key}")
        out[key] = value

    return out


def _write_json(path: Path, payload: dict[str, dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    path.write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="XLSM에서 문제 메타데이터 JSON 생성")
    parser.add_argument("--xlsm", required=True, help="입력 xlsm 파일 경로")
    parser.add_argument("--problems-sheet", default=PROBLEMS_SPEC.name, help="수능/모평/교육청 시트 이름")
    parser.add_argument("--essay-sheet", default=ESSAY_SPEC.name, help="논술 시트 이름")
    args = parser.parse_args()

    xlsm_path = Path(args.xlsm).expanduser().resolve()
    if not xlsm_path.is_file():
        raise FileNotFoundError(f"xlsm 파일을 찾을 수 없습니다: {xlsm_path}")

    wb = load_workbook(filename=xlsm_path, data_only=True, read_only=True, keep_vba=True)
    try:
        problems_spec = SheetSpec(
            name=args.problems_sheet,
            required_columns=PROBLEMS_SPEC.required_columns,
            output_path=PROBLEMS_SPEC.output_path,
        )
        essay_spec = SheetSpec(
            name=args.essay_sheet,
            required_columns=ESSAY_SPEC.required_columns,
            output_path=ESSAY_SPEC.output_path,
        )

        for spec in (problems_spec, essay_spec):
            if spec.name not in wb.sheetnames:
                raise ValueError(f"시트를 찾을 수 없습니다: {spec.name}")

        problems = _read_sheet(wb[problems_spec.name], SheetSpec("problems", problems_spec.required_columns, problems_spec.output_path))
        essays = _read_sheet(wb[essay_spec.name], SheetSpec("essay-problems", essay_spec.required_columns, essay_spec.output_path))
    finally:
        wb.close()

    _write_json(PROBLEMS_SPEC.output_path, problems)
    _write_json(ESSAY_SPEC.output_path, essays)

    print(f"완료: {PROBLEMS_SPEC.output_path}")
    print(f"완료: {ESSAY_SPEC.output_path}")
    print(f"problems: {len(problems)}건, essay-problems: {len(essays)}건")


if __name__ == "__main__":
    main()
